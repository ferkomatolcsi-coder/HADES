"""
HADES v2 - hades_export.py
Excel dashboard export with charts, color coding, and disk sheets
"""

import sys
import sqlite3
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

HADES_DIR = Path.home() / "Desktop" / "HADES"
DB_PATH = HADES_DIR / "hades.db"
EXPORT_PATH = HADES_DIR / f"HADES_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

def find_last_export():
    """Megkeresi a legutóbbi export fájlt."""
    exports = sorted(HADES_DIR.glob("HADES_export_*.xlsx"), reverse=True)
    # Az éppen most generálandót kihagyjuk
    for f in exports:
        if f != EXPORT_PATH:
            return f
    return None

def load_cached_sheets(last_export_path):
    """Betölti a korábbi export sheet snapshot-jait (label -> version_id)."""
    if not last_export_path:
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(last_export_path, read_only=True)
        cached = {}
        for sheet in wb.sheetnames:
            # A2 cellában van: "Scan: ... | N fájl | X GB"
            if sheet.startswith("💾 "):
                ws = wb[sheet]
                label = sheet.replace("💾 ", "").strip()
                a2 = ws["A2"].value or ""
                cached[label] = a2  # snapshot string
        wb.close()
        return cached
    except Exception as e:
        print(f"[HADES] Cache betöltés sikertelen: {e}")
        return {}

# --- Colors ---
C_HEADER_BG    = "1a1a2e"   # Sötétkék - fejléc háttér
C_HEADER_FG    = "FFFFFF"   # Fehér - fejléc szöveg
C_DASH_BG      = "16213e"   # Dashboard háttér
C_ACCENT       = "0f3460"   # Akcentus
C_PURPLE       = "7b2d8b"   # Javaslatok fejléc
C_RED          = "FF4444"   # 50GB+
C_ORANGE       = "FF8C00"   # 20GB+
C_GREEN        = "00AA44"   # OK
C_GRAY         = "CCCCCC"   # Üres/kis méret
C_ROW_ALT      = "F5F5F5"   # Alternáló sor

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(cell, bg=C_HEADER_BG, fg=C_HEADER_FG, size=11, bold=True):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def data_style(cell, bold=False, color="000000", bg=None, align="left"):
    cell.font = Font(name="Arial", bold=bold, color=color, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = thin_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)

def size_color(gb):
    if gb >= 50:   return C_RED
    if gb >= 20:   return C_ORANGE
    if gb >= 1:    return C_GREEN
    return C_GRAY

def format_gb(b):
    return round(b / (1024**3), 2)

def format_mb(b):
    return round(b / (1024**2), 2)

def load_disk_data():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Összes lemez
    c.execute("SELECT id, label, platform, last_seen FROM disks ORDER BY label")
    disks = c.fetchall()

    result = []
    for disk_id, label, platform, last_seen in disks:
        # Legfrissebb verzió
        c.execute("""
            SELECT id, scanned_at, file_count, total_bytes
            FROM scan_versions WHERE disk_id=?
            ORDER BY scanned_at DESC LIMIT 1
        """, (disk_id,))
        ver = c.fetchone()

        # Változás history
        c.execute("""
            SELECT logged_at, added, removed, modified
            FROM diff_log WHERE disk_id=?
            ORDER BY logged_at DESC LIMIT 10
        """, (disk_id,))
        diffs = c.fetchall()

        # Fájlok legfrissebb verzióból
        files = []
        if ver:
            c.execute("""
                SELECT path, size_bytes, modified_at
                FROM files WHERE version_id=?
                ORDER BY size_bytes DESC
            """, (ver[0],))
            files = c.fetchall()

        result.append({
            "id": disk_id,
            "label": label,
            "platform": platform,
            "last_seen": last_seen,
            "version": ver,
            "files": files,
            "diffs": diffs
        })

    conn.close()
    return result


def build_dashboard(wb, disks):
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False

    # Cím
    ws.merge_cells("A1:H1")
    ws["A1"] = "🔱 HADES – Hard Disk Explorer & Storage"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=16)
    ws["A1"].fill = PatternFill("solid", start_color=C_DASH_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Arial", color="AAAAAA", size=9, italic=True)
    ws["A2"].fill = PatternFill("solid", start_color=C_DASH_BG)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Lemez összefoglaló fejléc
    headers = ["💾 Lemez", "Platform", "Fájlok", "Méret (GB)", "Szabad (est.)", "Utoljára látva", "Státusz", "Sheet"]
    ws.row_dimensions[4].height = 28
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        header_style(cell)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    total_files = 0
    total_bytes = 0

    for i, disk in enumerate(disks, 5):
        ws.row_dimensions[i].height = 22
        ver = disk["version"]
        gb = format_gb(ver[3]) if ver else 0
        files = ver[2] if ver else 0
        total_files += files
        total_bytes += ver[3] if ver else 0

        col = size_color(gb)
        bg = C_ROW_ALT if i % 2 == 0 else "FFFFFF"

        vals = [
            disk["label"],
            disk["platform"] or "—",
            files,
            gb,
            "—",
            disk["last_seen"][:19] if disk["last_seen"] else "—",
            "✅ Aktív" if ver else "❌ Nem látott",
            f"→ {disk['label']}"
        ]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            data_style(cell, bg=bg if j != 4 else None)
            if j == 4:
                cell.fill = PatternFill("solid", start_color=col)
                cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Összesítő sor
    sum_row = len(disks) + 5
    ws.row_dimensions[sum_row].height = 24
    ws.merge_cells(f"A{sum_row}:B{sum_row}")
    ws[f"A{sum_row}"] = "ÖSSZESEN"
    header_style(ws[f"A{sum_row}"], bg=C_ACCENT)
    ws[f"C{sum_row}"] = total_files
    header_style(ws[f"C{sum_row}"], bg=C_ACCENT)
    ws[f"D{sum_row}"] = format_gb(total_bytes)
    header_style(ws[f"D{sum_row}"], bg=C_ACCENT)

    # Javaslatok szekció
    jav_row = sum_row + 2
    ws.merge_cells(f"A{jav_row}:H{jav_row}")
    ws[f"A{jav_row}"] = "💡 JAVASLATOK"
    header_style(ws[f"A{jav_row}"], bg=C_PURPLE, size=12)
    ws.row_dimensions[jav_row].height = 28

    suggestions = []
    for disk in disks:
        ver = disk["version"]
        if not ver:
            continue
        gb = format_gb(ver[3])
        if gb >= 50:
            suggestions.append(f"⚠️  {disk['label']}: {gb} GB – Erősen teli! Archiválás javasolt.")
        elif gb >= 20:
            suggestions.append(f"🟠 {disk['label']}: {gb} GB – Figyelembe venni.")

    if not suggestions:
        suggestions = ["✅ Minden lemez rendben – nincs kritikus figyelmeztetés."]

    for k, s in enumerate(suggestions, jav_row + 1):
        ws.merge_cells(f"A{k}:H{k}")
        ws[f"A{k}"] = s
        ws[f"A{k}"].font = Font(name="Arial", size=10, color="333333")
        ws[f"A{k}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws[f"A{k}"].border = thin_border()
        ws.row_dimensions[k].height = 20

    # --- PIE CHART ---
    chart_row = jav_row + len(suggestions) + 3

    # Data for pie chart (hidden area)
    data_start = chart_row
    ws[f"J{data_start}"] = "Lemez"
    ws[f"K{data_start}"] = "GB"
    for idx, disk in enumerate(disks, 1):
        ver = disk["version"]
        gb = format_gb(ver[3]) if ver else 0
        ws[f"J{data_start+idx}"] = disk["label"]
        ws[f"K{data_start+idx}"] = gb

    pie = PieChart()
    pie.title = "Lemezek méret szerint (GB)"
    pie.style = 10
    pie.width = 14
    pie.height = 10

    data_ref = Reference(ws, min_col=11, min_row=data_start, max_row=data_start+len(disks))
    labels_ref = Reference(ws, min_col=10, min_row=data_start+1, max_row=data_start+len(disks))
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    ws.add_chart(pie, f"A{chart_row}")

    # --- BAR CHART ---
    bar = BarChart()
    bar.type = "col"
    bar.title = "Fájlok száma lemezenként"
    bar.style = 10
    bar.width = 14
    bar.height = 10

    # Data for bar (hidden)
    bc_start = data_start + len(disks) + 2
    ws[f"J{bc_start}"] = "Lemez"
    ws[f"K{bc_start}"] = "Fájlok"
    for idx, disk in enumerate(disks, 1):
        ver = disk["version"]
        ws[f"J{bc_start+idx}"] = disk["label"]
        ws[f"K{bc_start+idx}"] = ver[2] if ver else 0

    data_ref2 = Reference(ws, min_col=11, min_row=bc_start, max_row=bc_start+len(disks))
    labels_ref2 = Reference(ws, min_col=10, min_row=bc_start+1, max_row=bc_start+len(disks))
    bar.add_data(data_ref2, titles_from_data=True)
    bar.set_categories(labels_ref2)
    ws.add_chart(bar, f"E{chart_row}")

    ws.freeze_panes = "A5"


def build_disk_sheet(wb, disk):
    label = disk["label"]
    ws = wb.create_sheet(title=f"💾 {label}")
    ws.sheet_view.showGridLines = False

    # Fejléc
    ws.merge_cells("A1:F1")
    ws["A1"] = f"💾 {label} – Fájl index"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", start_color=C_HEADER_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ver = disk["version"]
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Scan: {ver[1][:19] if ver else '—'} | {ver[2] if ver else 0} fájl | {format_gb(ver[3]) if ver else 0} GB"
    ws["A2"].font = Font(name="Arial", color="666666", size=9, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Oszlop fejlécek
    cols = ["#", "Fájlnév", "Mappa", "Méret (MB)", "Módosítva", "Kategória"]
    widths = [6, 35, 45, 14, 20, 14]
    ws.row_dimensions[4].height = 26
    for i, (h, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=4, column=i, value=h)
        header_style(cell)
        ws.column_dimensions[get_column_letter(i)].width = w

    # Fájlok
    for idx, (path, size, modified) in enumerate(disk["files"], 1):
        row = idx + 4
        ws.row_dimensions[row].height = 18
        bg = C_ROW_ALT if idx % 2 == 0 else "FFFFFF"

        p = Path(path)
        filename = p.name
        folder = str(p.parent)
        mb = format_mb(size)

        # Kategória
        ext = p.suffix.lower()
        if ext in [".jpg", ".jpeg", ".png", ".gif", ".heic", ".raw", ".cr2"]:
            cat = "📷 Kép"
        elif ext in [".mp4", ".mov", ".avi", ".mkv"]:
            cat = "🎬 Videó"
        elif ext in [".mp3", ".flac", ".wav", ".aac"]:
            cat = "🎵 Zene"
        elif ext in [".pdf", ".doc", ".docx", ".txt", ".pages"]:
            cat = "📄 Dokument"
        elif ext in [".zip", ".rar", ".tar", ".gz", ".7z"]:
            cat = "📦 Archív"
        elif ext in [".py", ".sh", ".js", ".ts", ".html", ".css"]:
            cat = "💻 Kód"
        else:
            cat = "📁 Egyéb"

        # Méret szín
        gb = mb / 1024
        col_bg = None
        col_fg = "000000"
        if gb >= 50:
            col_bg = C_RED
            col_fg = "FFFFFF"
        elif gb >= 20:
            col_bg = C_ORANGE
            col_fg = "FFFFFF"

        vals = [idx, filename, folder, mb, modified[:19] if modified else "—", cat]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            if j == 4 and col_bg:
                cell.fill = PatternFill("solid", start_color=col_bg)
                cell.font = Font(name="Arial", bold=True, color=col_fg, size=10)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = thin_border()
            else:
                data_style(cell, bg=bg, align="right" if j in [1, 4] else "left")

    # AutoFilter
    ws.auto_filter.ref = f"A4:F{len(disk['files'])+4}"
    ws.freeze_panes = "A5"


def build_history_sheet(wb, disks):
    ws = wb.create_sheet(title="📈 Változás history")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "📈 HADES – Változás history"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill("solid", start_color=C_DASH_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    headers = ["Lemez", "Időpont", "➕ Hozzáadva", "➖ Törölve", "🔄 Módosítva", "Összesen változás", "Trend"]
    widths = [16, 22, 16, 16, 16, 20, 14]
    ws.row_dimensions[3].height = 26
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=i, value=h)
        header_style(cell)
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 4
    for disk in disks:
        for logged_at, added, removed, modified in disk["diffs"]:
            ws.row_dimensions[row].height = 20
            bg = C_ROW_ALT if row % 2 == 0 else "FFFFFF"
            total = added + removed + modified
            trend = "📈 Növekedés" if added > removed else ("📉 Csökkenés" if removed > added else "➡️ Stabil")

            vals = [disk["label"], logged_at[:19], added, removed, modified, total, trend]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=j, value=v)
                data_style(cell, bg=bg, align="center" if j > 1 else "left")
            row += 1

    if row > 4:
        ws.auto_filter.ref = f"A3:G{row-1}"
    ws.freeze_panes = "A4"


def copy_sheet_from_workbook(src_path, src_title, dst_wb):
    """Átmásol egy sheetet a régi exportból az újba."""
    try:
        from openpyxl import load_workbook
        src_wb = load_workbook(src_path)
        if src_title not in src_wb.sheetnames:
            return False
        src_ws = src_wb[src_title]
        dst_ws = dst_wb.create_sheet(title=src_title)
        for row in src_ws.iter_rows():
            for cell in row:
                dst_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    dst_cell.font = cell.font.copy()
                    dst_cell.fill = cell.fill.copy()
                    dst_cell.border = cell.border.copy()
                    dst_cell.alignment = cell.alignment.copy()
        # Merged cells
        for merge in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(merge))
        # Column widths
        for col, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[col].width = dim.width
        # Row heights
        for row, dim in src_ws.row_dimensions.items():
            dst_ws.row_dimensions[row].height = dim.height
        # Freeze
        if src_ws.freeze_panes:
            dst_ws.freeze_panes = src_ws.freeze_panes
        # AutoFilter
        if src_ws.auto_filter.ref:
            dst_ws.auto_filter.ref = src_ws.auto_filter.ref
        src_wb.close()
        return True
    except Exception as e:
        print(f"[HADES] Sheet másolás hiba ({src_title}): {e}")
        return False


def export():
    print(f"\n[HADES] Loading database...")
    disks = load_disk_data()

    if not disks:
        print("[HADES] No data in database!")
        return

    print(f"[HADES] Found {len(disks)} disks")

    wb = Workbook()

    # Cache betöltés
    last_export = find_last_export()
    cached_sheets = load_cached_sheets(last_export)
    if last_export:
        print(f"[HADES] Cache: {last_export.name} ({len(cached_sheets)} cached sheet)")

    # 1. Dashboard
    print("[HADES] Building Dashboard...")
    build_dashboard(wb, disks)

    # 2. Lemezenként sheet - csak ha változott
    for disk in disks:
        label = disk["label"]
        ver = disk["version"]
        sheet_title = f"💾 {label}"

        # Aktuális snapshot string (ugyanolyan mint amit A2-be írunk)
        current_snap = f"Scan: {ver[1][:19] if ver else '—'} | {ver[2] if ver else 0} fájl | {format_gb(ver[3]) if ver else 0} GB"

        if label in cached_sheets and cached_sheets[label] == current_snap and last_export:
            # Nincs változás - régi sheet másolása
            print(f"[HADES] ♻️  Sheet unchanged, copying cache: {label}")
            if not copy_sheet_from_workbook(last_export, sheet_title, wb):
                print(f"[HADES] Cache másolás sikertelen, regenerálás: {label}")
                build_disk_sheet(wb, disk)
        else:
            print(f"[HADES] 🔄 Building sheet: {label} ({len(disk['files'])} files)...")
            build_disk_sheet(wb, disk)

    # 3. History
    print("[HADES] Building History sheet...")
    build_history_sheet(wb, disks)

    wb.save(EXPORT_PATH)
    print(f"\n[HADES] ✅ Export kész: {EXPORT_PATH}")
    print(f"[HADES] Méret: {EXPORT_PATH.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    export()
